import openpyxl
import cv2
import dlib
from os import listdir
from os.path import isfile, isdir, join
from imutils import face_utils
import numpy as np

mypath = './21'

cclemon = 1

files = listdir(mypath)

for f in files:
    fullpath = join(mypath, f)

    img = cv2.imread(fullpath)
        

    detector = dlib.get_frontal_face_detector()
    predictor = dlib.shape_predictor("../shape_predictor_106_face_landmarks.dat")
    # load the input image, resize it, and convert it to grayscale
    # image = cv2.imread('./img/warped_' + target_name)

    mask1 = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)

    mask = mask1.copy()


    # detect faces in the grayscale image
    rects = detector(mask1, 1)
    for (i, rect) in enumerate(rects):


        shape = predictor(mask1, rect)

        shape = face_utils.shape_to_np(shape)

        mask1 = face_utils.visualize_facial_landmarks(mask1, shape, True)

        cv2.imwrite("mask1.png", mask1)

        y_max = mask1.shape[0]
        x_max = mask1.shape[1]

        count = 0

        totalb = 0
        totalg = 0
        totalr = 0

        gray = []

        left_x = 0

        right_x = 0

        # cv2.imshow('', mask1)
        # cv2.waitKey(0)

        for y in range(y_max):
            for x in range(x_max):

                if mask1[x, y] != 0 and mask1[x, y] != 125:
                    gray.append(mask[x, y])

        gray = np.array(gray)

        y = np.array(gray)

        mean = int(np.mean(gray))

        std = int(np.std(gray, ddof=1))

        shadow = mean - 2 * std

        light = mean + 1.5 * std

        color = []

        # cv2.imshow('',mask)
        # cv2.waitKey(0)

        for i in range(512):
            for j in range(512):
                if mask1[i,j] == 255:
                    if mask[i,j] <shadow:
                        img[i,j,0] = 255
                        img[i,j,1] = 0
                        img[i,j,2] = 0
                    elif mask[i,j]>light:
                        img[i,j,0] = 0
                        img[i,j,1] = 0
                        img[i,j,2] = 255

        for c in gray:
            if shadow < c and c < light:
                color.append(c)

        color = np.array(color)

    wb = openpyxl.load_workbook('test.xlsx')

    wb.active = 5
    ws = wb.active

    
    ws.cell(column=1, row=cclemon).value = f
    ws.cell(column=2, row=cclemon).value = mean
    ws.cell(column=3, row=cclemon).value = int(np.mean(color))
    cclemon += 1


    # wb.save('test.xlsx')